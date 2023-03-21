import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

ws = load_workbook("res/input.xlsx")["Sheet1"]
max_row = ws.max_row

data = []
for i in range(1, max_row + 1):
    data.append([ws[f"A{i}"].value, ws[f"B{i}"].value])

data = np.array(data)
x = data[:, 0]
y = data[:, 1]

dy = y[1:] - y[:-1]
dx = x[1:] - x[:-1]

yprime = y.copy()
yprime[1:] = dy / dx

yprime[0] = yprime[1]

# d^2y / dx^2 = [y[i+2] - 2y[i+1] + y[i]] / [(x[i+1] - x[i])^2]
dpow2y = y.copy()
dpow2y = np.delete(dpow2y, 0)
dpow2y[1:] = (y[2:] - 2 * y[1:-1] + y[:-2])
dpow2y[0] = dpow2y[1]

dxpow2 = np.square(dx)

ydblPrime = dpow2y / dxpow2
ydblPrime = np.append(ydblPrime, ydblPrime[-1])

plt.plot(x, y, label="f(x) = (x^3)/2 + e^x + 4x - 6")
plt.plot(x, yprime, label="f’(x) = (3*(x^2))/2 + e^x + 4")
plt.plot(x, ydblPrime, label="f’’(x) = 3x + e^x")
plt.xlabel("x")
plt.ylabel("y")
plt.title("f(x) = (x^3)/2 + e^x + 4x - 6")
plt.legend()
plt.savefig('res/plot.png')

wb_save = Workbook()
ws_save = wb_save.active

i = 0 
for x_i in x:
    ws_save[f"A{i+1}"] = x_i
    ws_save[f"B{i+1}"] = yprime[i]
    ws_save[f"C{i+1}"] = ydblPrime[i]
    i += 1

wb_save.save("res/output.xlsx")
